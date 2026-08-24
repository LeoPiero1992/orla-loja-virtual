"""Atualiza o estoque da ORLA pelo SharePoint ou pela cópia local sincronizada."""

from __future__ import annotations

import base64
import json
import os
import re
import tempfile
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SHEET_NAME = "ProntaEntrega"
BALANCE_HEADER = "Saldo Disponível"


def required(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Variável obrigatória ausente: {name}")
    return value


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return re.sub(r"[^a-z0-9]+", " ", "".join(ch for ch in text if unicodedata.category(ch) != "Mn").lower()).strip()


def graph_token() -> str:
    tenant = required("SP_TENANT_ID")
    payload = urllib.parse.urlencode({
        "client_id": required("SP_CLIENT_ID"),
        "client_secret": required("SP_CLIENT_SECRET"),
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }).encode()
    request = urllib.request.Request(
        f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
        data=payload,
        headers={"content-type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        return json.load(response)["access_token"]


def share_id(url: str) -> str:
    encoded = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")
    return f"u!{encoded}"


def download_workbook(destination: Path) -> None:
    url = required("SHAREPOINT_FILE_URL")
    request = urllib.request.Request(
        f"https://graph.microsoft.com/v1.0/shares/{share_id(url)}/driveItem/content",
        headers={"authorization": f"Bearer {graph_token()}"},
    )
    with urllib.request.urlopen(request, timeout=120) as response, destination.open("wb") as output:
        while chunk := response.read(1024 * 1024):
            output.write(chunk)


def find_header(headers: list[object], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if normalized(header) in candidates:
            return index
    return None


def find_balance_header(headers: list[object]) -> int | None:
    for index, header in enumerate(headers):
        value = normalized(header)
        if value == normalized(BALANCE_HEADER) or value.startswith("saldo dispon"):
            return index
    return None


def digits(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value or ""))


def clean_sku(value: object) -> str:
    text = str(value or "").strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"[^0-9A-Z]", "", text)


def build_sku(row: tuple[object, ...], indexes: dict[str, int | None]) -> str:
    if indexes["sku"] is not None:
        direct = clean_sku(row[indexes["sku"]])
        if direct:
            return direct
    reference = digits(row[indexes["reference"]]) if indexes["reference"] is not None else ""
    catalog = digits(row[indexes["catalog"]]) if indexes["catalog"] is not None else ""
    color = digits(row[indexes["color"]]) if indexes["color"] is not None else ""
    size = clean_sku(row[indexes["size"]]) if indexes["size"] is not None else ""
    if len(reference) == 3 and catalog:
        reference += catalog.zfill(2)[-2:]
    if not reference or not color or not size:
        return ""
    return f"{reference}{color.zfill(3)[-3:]}{size}"


def extract_stock(workbook_path: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(f"Aba '{SHEET_NAME}' não encontrada. Abas disponíveis: {', '.join(workbook.sheetnames)}")
        sheet = workbook[SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        headers: list[object] = []
        for candidate in rows:
            values = list(candidate)
            if find_header(values, {"sku"}) is not None and find_balance_header(values) is not None:
                headers = values
                break
        if not headers:
            raise RuntimeError(
                f"Não foi encontrada na aba '{SHEET_NAME}' uma linha contendo SKU e '{BALANCE_HEADER}'."
            )
        normalized_headers = [normalized(value) for value in headers]
        balance = find_balance_header(headers)
        if balance is None:
            raise RuntimeError(f"Coluna '{BALANCE_HEADER}' não encontrada. Cabeçalhos: {normalized_headers}")
        indexes = {
            "sku": find_header(headers, {"sku", "codigo sku", "cod sku", "codigo de barras"}),
            "reference": find_header(headers, {"referencia", "ref", "codigo produto", "cod produto"}),
            "catalog": find_header(headers, {"catalogo", "colecao", "numero catalogo", "n catalogo"}),
            "color": find_header(headers, {"cor", "codigo cor", "cod cor"}),
            "size": find_header(headers, {"grade", "tamanho", "tam"}),
        }
        if indexes["sku"] is None and any(indexes[key] is None for key in ("reference", "color", "size")):
            raise RuntimeError("Não foi possível identificar o SKU nem as colunas Referência, Cor e Grade.")
        stock: dict[str, int] = {}
        for row in rows:
            sku = build_sku(row, indexes)
            if not sku:
                continue
            try:
                quantity = max(0, int(float(row[balance] or 0)))
            except (TypeError, ValueError):
                quantity = 0
            stock[sku] = stock.get(sku, 0) + quantity
        if not stock:
            raise RuntimeError("Nenhum SKU válido foi encontrado na aba ProntaEntrega.")
        return dict(sorted(stock.items()))
    finally:
        workbook.close()


def write_outputs(stock: dict[str, int], source: str) -> None:
    synced_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    stock_js = "window.ORLA_STORE_STOCK=" + json.dumps(stock, ensure_ascii=False, separators=(",", ":")) + ";\n"
    meta = json.dumps({"source": source, "syncedAt": synced_at}, ensure_ascii=False, indent=2) + "\n"
    for relative, content in (
        ("stock-data.js", stock_js),
        ("stock-meta.json", meta),
        ("dist/client/stock-data.js", stock_js),
        ("dist/client/stock-meta.json", meta),
    ):
        path = ROOT / relative
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        temporary.write_text(content, encoding="utf-8", newline="\n")
        temporary.replace(path)
    print(f"Estoque atualizado: {len(stock)} SKUs, sincronizado em {synced_at}")


def main() -> None:
    local_file = os.environ.get("LOCAL_STOCK_FILE", "").strip()
    if local_file:
        workbook_path = Path(local_file).expanduser().resolve()
        if not workbook_path.is_file():
            raise RuntimeError(f"Planilha local não encontrada: {workbook_path}")
        write_outputs(
            extract_stock(workbook_path),
            "OneDrive local/ProntaEntrega/Saldo Disponível",
        )
        return
    with tempfile.TemporaryDirectory() as directory:
        workbook_path = Path(directory) / "pronta-entrega.xlsm"
        download_workbook(workbook_path)
        write_outputs(
            extract_stock(workbook_path),
            "SharePoint/ProntaEntrega/Saldo Disponível",
        )


if __name__ == "__main__":
    main()
